import os
import uuid
from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg

from io import BytesIO

from .ocr import extract_data
from .models import VerifikasiIjazah
from .metrics import compute_cer, compute_wer


# ======================
# HELPER: SIDEBAR STATS
# ======================
def _sidebar_context():
    all_data = VerifikasiIjazah.objects.all()
    return {
        "total": all_data.count(),
        "valid": all_data.filter(status__iexact="VALID").count(),
        "pending": all_data.filter(status__icontains="MENUNGGU").count(),
    }


# ======================
# BACKGROUND PROCESS
# ======================
def process_file_background(obj_id, file_path, original_filename):
    try:
        nama_ocr, tahun, hasil_ocr = extract_data(file_path)
        obj = VerifikasiIjazah.objects.get(id=obj_id)

        if nama_ocr and nama_ocr != "Perlu Verifikasi Manual":
            obj.nama = nama_ocr

        obj.extracted_year = str(tahun) if tahun else ""

        if not tahun:
            obj.status = "TIDAK TERDETEKSI"
        else:
            obj.status = "MENUNGGU VERIFIKASI"

        obj.save()
        print("✔ DONE:", original_filename, "|", obj.nama, "|", tahun)

    except Exception as e:
        print("ERROR:", e)
        try:
            obj = VerifikasiIjazah.objects.get(id=obj_id)
            obj.status = "TIDAK TERDETEKSI"
            obj.save()
        except:
            pass


# ======================
# UPLOAD
# ======================
@login_required(login_url="login")
def upload_ijazah(request):
    hasil_ocr = ""
    uploaded_files = []

    last_data = VerifikasiIjazah.objects.order_by("-created_at").first()

    if request.method == "GET" and request.GET.get("reset") == "1":
        request.session.flush()

    if request.method == "POST" and request.FILES.getlist("ijazah"):
        files = request.FILES.getlist("ijazah")

        for file in files:
            os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
            filename = f"{uuid.uuid4()}_{file.name}"
            file_path = os.path.join(settings.MEDIA_ROOT, filename)

            with open(file_path, "wb+") as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            file_lower = file.name.lower()
            is_image = file_lower.endswith((".png", ".jpg", ".jpeg"))
            is_pdf = file_lower.endswith(".pdf")

            print("START PROCESS:", file.name)
            nama_ocr, tahun, hasil_ocr = extract_data(file_path)

            if (
                not nama_ocr
                or nama_ocr == "Perlu Verifikasi Manual"
                or nama_ocr == "Tidak Terdeteksi"
            ):
                nama_final = os.path.splitext(file.name)[0]
            else:
                nama_final = nama_ocr

            if not tahun:
                status = "TIDAK TERDETEKSI"
                hasil_benar = False
                salah_tahun = False
                tidak_terbaca = True
            else:
                status = "MENUNGGU VERIFIKASI"
                hasil_benar = True
                salah_tahun = False
                tidak_terbaca = False

            file.seek(0)

            obj = VerifikasiIjazah.objects.create(
                nama=nama_final,
                nama_ocr=nama_final,
                nim="-",
                file=file,
                extracted_year=str(tahun) if tahun else "",
                status=status,
                hasil_benar=hasil_benar,
                salah_tahun=salah_tahun,
                tidak_terbaca=tidak_terbaca,
            )

            print("✔ DONE:", file.name, "|", obj.nama, "|", tahun)

            uploaded_files.append({
                "nama": obj.nama,
                "tahun": obj.extracted_year or "-",
                "file_url": obj.file.url if obj.file else None,
                "is_image": is_image,
                "is_pdf": is_pdf,
            })

        last_data = VerifikasiIjazah.objects.order_by("-created_at").first()

    return render(request, "upload_admin.html", {
        "data": last_data,
        "uploaded_files": uploaded_files,
        **_sidebar_context(),
    })


# ======================
# UPLOAD SINGLE (AJAX)
# ======================
@login_required(login_url="login")
def upload_single(request):
    if request.method != "POST" or not request.FILES.get("ijazah"):
        return JsonResponse({"error": "No file"}, status=400)

    file = request.FILES["ijazah"]

    import tempfile
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.name)

    with open(file_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    file_lower = file.name.lower()
    is_image = file_lower.endswith((".png", ".jpg", ".jpeg"))
    is_pdf = file_lower.endswith(".pdf")

    print("START PROCESS:", file.name)
    nama_ocr, tahun, hasil_ocr = extract_data(file_path)

    if (
        not nama_ocr
        or nama_ocr == "Perlu Verifikasi Manual"
        or nama_ocr == "Tidak Terdeteksi"
    ):
        nama_final = os.path.splitext(file.name)[0]
    else:
        nama_final = nama_ocr

    if not tahun:
        status = "TIDAK TERDETEKSI"
        hasil_benar = False
        salah_tahun = False
        tidak_terbaca = True
    else:
        status = "MENUNGGU VERIFIKASI"
        hasil_benar = True
        salah_tahun = False
        tidak_terbaca = False

    file.seek(0)

    obj = VerifikasiIjazah.objects.create(
        nama=nama_final,
        nama_ocr=nama_final,
        nim="-",
        file=file,
        extracted_year=str(tahun) if tahun else "",
        status=status,
        hasil_benar=hasil_benar,
        salah_tahun=salah_tahun,
        tidak_terbaca=tidak_terbaca,
    )

    print("✔ DONE:", file.name, "|", obj.nama, "|", tahun)

    file_url = obj.file.url if obj.file else None
    file_url = fix_cloudinary_url(file_url, obj.file.name if obj.file else None)

    return JsonResponse({
        "nama": obj.nama,
        "tahun": obj.extracted_year or "-",
        "file_url": file_url,
        "is_image": is_image,
        "is_pdf": is_pdf,
    })

# ======================
# HITUNG CER/WER
# ======================
def _update_metrics(obj):
    if obj.nama_ocr:
        obj.cer = round(compute_cer(obj.nama_ocr, obj.nama), 4)
        obj.wer = round(compute_wer(obj.nama_ocr, obj.nama), 4)
    else:
        obj.cer = 0.0
        obj.wer = 0.0


def _get_avg_metrics(queryset=None):
    if queryset is None:
        queryset = VerifikasiIjazah.objects.all()

    evaluated = queryset.filter(
        cer__isnull=False,
        wer__isnull=False,
        status__in=["VALID", "TIDAK MEMENUHI SYARAT"],
    )

    avg = evaluated.aggregate(avg_cer=Avg("cer"), avg_wer=Avg("wer"))
    avg_cer = avg["avg_cer"] or 0
    avg_wer = avg["avg_wer"] or 0

    return {
        "avg_cer": round(avg_cer * 100, 2),
        "avg_wer": round(avg_wer * 100, 2),
        "akurasi_cer": round((1 - avg_cer) * 100, 2),
        "akurasi_wer": round((1 - avg_wer) * 100, 2),
        "total_evaluated": evaluated.count(),
    }


# ======================
# DASHBOARD ADMIN
# ======================
@login_required(login_url="login")
def dashboard_admin(request):
    data_list = VerifikasiIjazah.objects.order_by("-created_at")

    search = request.GET.get("search")
    status = request.GET.get("status")

    if search:
        data_list = data_list.filter(nama__icontains=search)
    if status:
        if status == "VALID":
            data_list = data_list.filter(status__iexact="VALID")
        elif status == "MENUNGGU":
            data_list = data_list.filter(status__icontains="MENUNGGU")
        elif status == "TIDAK MEMENUHI SYARAT":
            data_list = data_list.filter(status__iexact="TIDAK MEMENUHI SYARAT")
        elif status == "TIDAK TERDETEKSI":
            data_list = data_list.filter(status__iexact="TIDAK TERDETEKSI")

    paginator = Paginator(data_list, 10)
    page_number = request.GET.get("page")
    data = paginator.get_page(page_number)

    all_data = VerifikasiIjazah.objects.all()
    metrics = _get_avg_metrics(all_data)

    total_count = all_data.count()
    valid_count = all_data.filter(status__iexact="VALID").count()
    accuracy = round((valid_count / total_count) * 100, 2) if total_count else 0

    return render(request, "dashboard_admin.html", {
        "data": data,
        "total": total_count,
        "valid": valid_count,
        "not_valid": all_data.filter(status__iexact="TIDAK MEMENUHI SYARAT").count(),
        "pending": all_data.filter(status__icontains="MENUNGGU").count(),
        "unknown_year": all_data.filter(status__iexact="TIDAK TERDETEKSI").count(),
        "today": all_data.filter(created_at__date=timezone.localdate()).count(),
        "accuracy": accuracy,
        "search": search or "",
        "status": status or "",
        **metrics,
    })


# ======================
# VERIFIKASI ADMIN
# ======================
@login_required(login_url="login")
def verifikasi_valid(request, id):
    data = get_object_or_404(VerifikasiIjazah, id=id)
    data.status = "VALID"
    _update_metrics(data)
    data.save()
    page = request.GET.get("page", "1")
    return redirect(f"/dashboard/?page={page}")


@login_required(login_url="login")
def verifikasi_tidak_sesuai(request, id):
    data = get_object_or_404(VerifikasiIjazah, id=id)
    data.status = "TIDAK MEMENUHI SYARAT"
    _update_metrics(data)
    data.save()
    page = request.GET.get("page", "1")
    return redirect(f"/dashboard/?page={page}")

# ======================
# REPORTS
# ======================
def reports_admin(request):
    context = _build_report_data(request)
    return render(request, "reports_admin.html", context)


def _build_report_data(request):
    today = timezone.localdate()
    data = VerifikasiIjazah.objects.order_by("-created_at")

    search = request.GET.get("search")
    status = request.GET.get("status")

    if search:
        data = data.filter(nama__icontains=search)
    if status:
        if status == "VALID":
            data = data.filter(status__iexact="VALID")
        elif status == "MENUNGGU":
            data = data.filter(status__icontains="MENUNGGU")
        elif status == "TIDAK MEMENUHI SYARAT":
            data = data.filter(status__iexact="TIDAK MEMENUHI SYARAT")
        elif status == "TIDAK TERDETEKSI":
            data = data.filter(status__iexact="TIDAK TERDETEKSI")

    paginator = Paginator(data, 10)
    page_number = request.GET.get("page")
    report_rows = paginator.get_page(page_number)

    import datetime
    start_of_month = timezone.make_aware(
        datetime.datetime(today.year, today.month, 1)
    )
    if today.month == 12:
        end_of_month = timezone.make_aware(
            datetime.datetime(today.year + 1, 1, 1)
        )
    else:
        end_of_month = timezone.make_aware(
            datetime.datetime(today.year, today.month + 1, 1)
        )
    
    month_data = VerifikasiIjazah.objects.filter(
        created_at__gte=start_of_month,
        created_at__lt=end_of_month,
    )

    total_month = month_data.count()
    valid = month_data.filter(status__iexact="VALID").count()
    not_valid = month_data.filter(status__iexact="TIDAK MEMENUHI SYARAT").count()
    pending = month_data.filter(status__icontains="MENUNGGU").count()
    unknown_year = month_data.filter(status__iexact="TIDAK TERDETEKSI").count()

    accuracy = round((valid / total_month) * 100, 2) if total_month else 0

    total_all = data.count()
    akurasi_total = round(
        (data.filter(status__iexact="VALID").count() / total_all) * 100, 2
    ) if total_all else 0

    all_data = VerifikasiIjazah.objects.all()
    metrics = _get_avg_metrics(data)

    return {
        "total_month": total_month,
        "valid": valid,
        "not_valid": not_valid,
        "pending": pending,
        "unknown_year": unknown_year,
        "accuracy": accuracy,
        "report_rows": report_rows,
        "akurasi_total": akurasi_total,
        "search": search or "",
        "status": status or "",
        "total": all_data.count(),
        **metrics,
    }


# ======================
# EXPORT EXCEL (XLSX)
# ======================
def download_reports_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    report_data = _build_report_data(request)
    all_rows = VerifikasiIjazah.objects.order_by("-created_at")

    search = request.GET.get("search")
    status_filter = request.GET.get("status")
    if search:
        all_rows = all_rows.filter(nama__icontains=search)
    if status_filter:
        if status_filter == "VALID":
            all_rows = all_rows.filter(status__iexact="VALID")
        elif status_filter == "MENUNGGU":
            all_rows = all_rows.filter(status__icontains="MENUNGGU")
        elif status_filter == "TIDAK MEMENUHI SYARAT":
            all_rows = all_rows.filter(status__iexact="TIDAK MEMENUHI SYARAT")
        elif status_filter == "TIDAK TERDETEKSI":
            all_rows = all_rows.filter(status__iexact="TIDAK TERDETEKSI")

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Verifikasi"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10B8CC", end_color="10B8CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "LAPORAN VERIFIKASI IJAZAH - REGMABA UNSRAT"
    title_cell.font = Font(bold=True, size=14, color="10B8CC")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Tanggal: {timezone.localdate().strftime('%d-%m-%Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(size=10, italic=True)

    ws["A4"].value = "RINGKASAN"
    ws["A4"].font = Font(bold=True, size=11)

    summary_labels = ["Total", "Valid", "Tidak Valid", "Menunggu", "Tidak Terdeteksi",
                       "CER Rata-rata", "WER Rata-rata", "Akurasi CER", "Akurasi WER"]
    summary_values = [
        report_data["total_month"], report_data["valid"], report_data["not_valid"],
        report_data["pending"], report_data["unknown_year"],
        f"{report_data['avg_cer']}%", f"{report_data['avg_wer']}%",
        f"{report_data['akurasi_cer']}%", f"{report_data['akurasi_wer']}%",
    ]

    for i, (label, value) in enumerate(zip(summary_labels, summary_values)):
        ws[f"A{5 + i}"].value = label
        ws[f"A{5 + i}"].font = Font(size=10)
        ws[f"B{5 + i}"].value = value
        ws[f"B{5 + i}"].font = Font(bold=True, size=10)

    start_row = 16
    headers = ["No", "Nama", "Tahun", "Status", "CER (%)", "WER (%)", "Tanggal"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, item in enumerate(all_rows, 1):
        row = start_row + i
        values = [
            i, item.nama,
            item.extracted_year or "-", item.status,
            round(item.cer * 100, 2) if item.cer is not None else "-",
            round(item.wer * 100, 2) if item.wer is not None else "-",
            item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "-",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align if col != 2 else Alignment(vertical="center")

    col_widths = [6, 35, 10, 25, 12, 12, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="laporan_verifikasi.xlsx"'
    wb.save(response)
    return response


# ======================
# EXPORT PDF
# ======================
def download_reports_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    report_data = _build_report_data(request)
    all_rows = VerifikasiIjazah.objects.order_by("-created_at")

    search = request.GET.get("search")
    status_filter = request.GET.get("status")
    if search:
        all_rows = all_rows.filter(nama__icontains=search)
    if status_filter:
        if status_filter == "VALID":
            all_rows = all_rows.filter(status__iexact="VALID")
        elif status_filter == "MENUNGGU":
            all_rows = all_rows.filter(status__icontains="MENUNGGU")
        elif status_filter == "TIDAK MEMENUHI SYARAT":
            all_rows = all_rows.filter(status__iexact="TIDAK MEMENUHI SYARAT")
        elif status_filter == "TIDAK TERDETEKSI":
            all_rows = all_rows.filter(status__iexact="TIDAK TERDETEKSI")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("CustomTitle", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#10B8CC"), alignment=1, spaceAfter=6)
    elements.append(Paragraph("LAPORAN VERIFIKASI IJAZAH", title_style))
    elements.append(Paragraph("REGMABA UNSRAT", title_style))

    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"],
        fontSize=10, alignment=1, textColor=colors.grey, spaceAfter=20)
    tanggal = timezone.localdate().strftime("%d-%m-%Y")
    elements.append(Paragraph(f"Tanggal: {tanggal}", subtitle_style))

    summary_data = [
        ["Total Bulan Ini", str(report_data["total_month"]),
         "Valid", str(report_data["valid"]),
         "Tidak Valid", str(report_data["not_valid"])],
        ["Menunggu", str(report_data["pending"]),
         "Tidak Terdeteksi", str(report_data["unknown_year"]),
         "", ""],
        ["CER Rata-rata", f"{report_data['avg_cer']}%",
         "WER Rata-rata", f"{report_data['avg_wer']}%",
         "Akurasi CER", f"{report_data['akurasi_cer']}%"],
    ]

    summary_table = Table(summary_data, colWidths=[4*cm, 3*cm, 4*cm, 3*cm, 4*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E3F7FA")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E3F7FA")),
        ("BACKGROUND", (4, 0), (4, -1), colors.HexColor("#E3F7FA")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("FONTNAME", (5, 0), (5, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)
    table_data = [["No", "Nama", "Tahun", "Status", "CER (%)", "WER (%)", "Tanggal"]]

    for i, item in enumerate(all_rows, 1):
        cer_val = f"{item.cer * 100:.2f}" if item.cer is not None else "-"
        wer_val = f"{item.wer * 100:.2f}" if item.wer is not None else "-"
        table_data.append([
            str(i),
            Paragraph(item.nama, cell_style),
            item.extracted_year or "-",
            item.status,
            cer_val, wer_val,
            item.created_at.strftime("%Y-%m-%d") if item.created_at else "-",
        ])

    col_widths = [1.2*cm, 7*cm, 2.5*cm, 4.5*cm, 2.5*cm, 2.5*cm, 3*cm]
    data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    data_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10B8CC")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGNMENT", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGNMENT", (0, 1), (0, -1), "CENTER"),
        ("ALIGNMENT", (2, 1), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(data_table)
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="laporan_verifikasi.pdf"'
    return response


# ======================
# VIEW DOCUMENT
# ======================
def view_document(request, document_id):
    item = get_object_or_404(VerifikasiIjazah, id=document_id)
    all_data = VerifikasiIjazah.objects.all()

    file_url = item.file.url if item.file else None
    file_url = fix_cloudinary_url(file_url, item.file.name if item.file else None)

    page = request.GET.get("page", "1")

    return render(request, "view_document.html", {
        "item": item,
        "file_url": file_url,
        "page": page,
        "total": all_data.count(),
        "valid": all_data.filter(status__iexact="VALID").count(),
        "pending": all_data.filter(status__icontains="MENUNGGU").count(),
    })


# ======================
# EDIT VERIFIKASI
# ======================
def edit_verifikasi(request):
    if request.method == "POST":
        item_id = request.POST.get("id")
        nama = request.POST.get("nama")
        tahun = request.POST.get("tahun")
        status = request.POST.get("status")
        page = request.POST.get("page", "1")

        item = VerifikasiIjazah.objects.get(id=item_id)
        item.nama = nama
        item.extracted_year = tahun
        if status:
            item.status = status

        _update_metrics(item)
        item.save()

        return redirect(f"/dashboard/?page={page}")

    return redirect("dashboard_admin")


# ======================
# CLOUDINARY
# ======================
def fix_cloudinary_url(url, filename):
    if not url:
        return url

    is_pdf = False
    
    if filename and filename.lower().endswith('.pdf'):
        is_pdf = True
    elif '/image/upload/' in url:
        image_exts = ('.jpg', '.jpeg', '.png', '.gif', '.webp')
        if not any(url.lower().endswith(ext) for ext in image_exts):
            is_pdf = True
    
    if is_pdf and '/image/upload/' in url:
        return url.replace('/image/upload/', '/image/upload/pg_1/') + '.jpg'
    
    return url


# ======================
# LOGIN LOGOUT
# ======================
def login_admin(request):
    error = None

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("dashboard_admin")
        else:
            error = "Username atau password salah."

    return render(request, "login.html", {"error": error})


def logout_admin(request):
    logout(request)
    return redirect("login")

# ======================
# KELOLA USER (SUPER ADMIN ONLY)
# ======================
@login_required(login_url="login")
def manage_users(request):
    if not request.user.is_superuser:
        return redirect("dashboard_admin")

    from django.contrib.auth.models import User
    users = User.objects.all().order_by("-date_joined")

    return render(request, "manage_users.html", {
        "users": users,
        **_sidebar_context(),
    })


@login_required(login_url="login")
def add_user(request):
    if not request.user.is_superuser:
        return redirect("dashboard_admin")

    if request.method == "POST":
        from django.contrib.auth.models import User

        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        role = request.POST.get("role", "operator")

        if not username or not password:
            return render(request, "manage_users.html", {
                "users": User.objects.all().order_by("-date_joined"),
                "error": "Username dan password wajib diisi.",
                "show_add_modal": True,
                **_sidebar_context(),
            })

        if User.objects.filter(username=username).exists():
            return render(request, "manage_users.html", {
                "users": User.objects.all().order_by("-date_joined"),
                "error": "Username sudah digunakan.",
                "show_add_modal": True,
                **_sidebar_context(),
            })

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
        )

        if role == "superadmin":
            user.is_superuser = True
            user.is_staff = True
            user.save()

        return redirect("manage_users")

    return redirect("manage_users")


@login_required(login_url="login")
def edit_user(request, user_id):
    if not request.user.is_superuser:
        return redirect("dashboard_admin")

    from django.contrib.auth.models import User
    target_user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        role = request.POST.get("role", "operator")

        if username:
            if User.objects.filter(username=username).exclude(id=target_user.id).exists():
                return render(request, "manage_users.html", {
                    "users": User.objects.all().order_by("-date_joined"),
                    "error": "Username sudah digunakan.",
                    **_sidebar_context(),
                })
            target_user.username = username

        if email:
            target_user.email = email

        if password:
            target_user.set_password(password)

        target_user.is_superuser = (role == "superadmin")
        target_user.is_staff = (role == "superadmin")
        target_user.save()

        return redirect("manage_users")

    return redirect("manage_users")


@login_required(login_url="login")
def delete_user(request, user_id):
    if not request.user.is_superuser:
        return redirect("dashboard_admin")

    from django.contrib.auth.models import User
    target_user = get_object_or_404(User, id=user_id)

    if target_user.id == request.user.id:
        return redirect("manage_users")

    target_user.delete()
    return redirect("manage_users")

# ======================
# SETTINGS
# ======================
@login_required(login_url="login")
def settings_admin(request):
    success = None
    error = None

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()

        user = request.user

        try:
            if username:
                from django.contrib.auth.models import User
                if User.objects.filter(username=username).exclude(id=user.id).exists():
                    error = "Username sudah digunakan."
                else:
                    user.username = username

            if email:
                user.email = email

            if password:
                user.set_password(password)

            if not error:
                user.save()
                success = "Profil berhasil diperbarui."

                if password:
                    from django.contrib.auth import update_session_auth_hash
                    update_session_auth_hash(request, user)

        except Exception as e:
            error = f"Gagal menyimpan: {str(e)}"

    all_data = VerifikasiIjazah.objects.all()

    return render(request, "settings_admin.html", {
        "admin": {
            "name": request.user.get_full_name() or request.user.username,
            "email": request.user.email,
            "role": "Super Admin" if request.user.is_superuser else "Admin",
        },
        "total": all_data.count(),
        "valid": all_data.filter(status__iexact="VALID").count(),
        "pending": all_data.filter(status__icontains="MENUNGGU").count(),
        "success": success,
        "error": error,
    })